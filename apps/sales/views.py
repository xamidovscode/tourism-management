import os
from datetime import datetime
import qrcode
from django.shortcuts import render
from django.template.loader import render_to_string
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.utils.cell import get_column_letter
from openpyxl.workbook.workbook import Workbook
from weasyprint import HTML
from core.settings.base import BASE_DIR
from .models import Sale
from django.http import HttpResponse
from .models import SoldTours


def sale_list(request):
    sales = Sale.objects.all()
    return render(request, 'sales/sale_list.html', {'sales': sales})


def export_pdf1(request, pk):
    sale = SoldTours.objects.get(pk=pk)
    now = datetime.now()

    qr_data = f"Receipt ID: {sale.pk}, Tour: {sale.tour}, Agent: {sale.agent}"
    qr_img = qrcode.make(qr_data)
    qr_path = f"static/qr/qr_{sale.pk}.png"
    full_qr_path = os.path.join(BASE_DIR, qr_path)
    os.makedirs(os.path.dirname(full_qr_path), exist_ok=True)
    qr_img.save(full_qr_path)

    context = {
        'sale': sale,
        'current_date': now.strftime('%d-%m-%Y %H:%M'),
        'processed_date': sale.processed_at.strftime('%d-%m-%Y') if sale.processed_at else '-',
        'qr_code_url': os.path.join(BASE_DIR, qr_path)
    }

    html_string = render_to_string('admin/customers/pdf_recaipt.html', context)
    html = HTML(string=html_string)
    pdf_file = html.write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="SoldTour_{sale.pk}.pdf"'
    return response


def export_pdf2(request, pk):
    sale = SoldTours.objects.get(pk=pk)
    now = datetime.now()

    context = {
        'sale': sale,
        'generated_at': now.strftime('%Y-%m-%d %H:%M'),
        'tour_name': sale.tour.name if sale.tour else '—',
        'tour_date': sale.tour.start_sale.strftime('%d-%m-%Y') if sale.tour and sale.tour.start_sale else '—',

        'hotel_name': sale.tour.hotel.name if sale.tour and sale.tour.hotel else 'Clifton International',
        'supplier': 'MB Safari' if sale.tour.supplier else '—',
        'vehicle': sale.tour.transfer_type.name if sale.tour and sale.tour.transfer_type else 'Motorcycle',

        'pax': {
            'adult': 0,
            'child': 0,
            'toddler': 0,
            'infant': 0,
        },

        'operator': sale.agent.full_name if sale.agent else '—',
        'notes': sale.description or 'No additional notes',
        'meeting_point': 'Hotel Reception',
        'guide_name': 'Stanislav Gorodilov',
        'excursion': sale.tour.type.name if sale.tour and sale.tour.type else '—',
    }

    html_string = render_to_string('admin/customers/pdf_invoice.html', context)
    html = HTML(string=html_string)
    pdf_file = html.write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Voucher_{sale.pk}.pdf"'
    return response


def export_excel(request, pk):
    sale = SoldTours.objects.get(pk=pk)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sold Tour"
    headers = ["Field", "Value"]
    data = [
        ["ID", str(sale.pk)],
        ["Created At", str(sale.created_at)],
        ["Tour Name", str(sale.tour)],
        ["Tour Date", str(sale.created_at)],
        ["Agent", str(sale.agent)],
        ["Pickup Time", str(sale.pick_up_time)],
        ["Processed At", str(sale.processed_at)],
        ["Description", sale.description or "-"],
        ["Area", str(sale.area or "-")],
        ["Adult", str(sale.tour or "-")],
        ["Child", str(sale.tour or "-")],
        ["Toodle", str(sale.tour or "-")],
        ["Discount", str(sale.discount)],
        ["Discount Type", str(sale.discount_type or 0)],
    ]

    ws.append(headers)
    for row in data:
        ws.append([str(item) for item in row])

    for cell in ws["1:1"]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col in range(1, 3):
        max_length = 0
        col_letter = get_column_letter(col)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='left')
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="soldtour_{sale.pk}.xlsx"'
    wb.save(response)
    # return response


